# -*- coding: UTF-8 -*-
import requests
import time
import re

from PIL import Image
#from pytesseract import *
import PIL.ImageOps
import hmac
import hashlib
import base64
import random

from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
from bs4 import BeautifulSoup as bs
import os
#from lxml import etree

import xlrd
import pymysql
import math
import sys
import request
reload(sys)
sys.setdefaultencoding("utf-8")

#识别验证码图片
def yanzheng_local():
    # 图像处理
    im = Image.open("captcha.png")
    im = im.convert('L')
    #im.show()
    im2 = im.point(lambda x: 0 if x > 200 else 255)
    #im2.show()
    im3 = im2.save("yanzhengma.png")

    # 腾讯免费接口识别
    appid = "1257122374"  # 腾讯云号码
    # bucket = "你的bucket"  # 不要也可以
    secret_id = "AKIDGK8gy23WxIZeU5lWLtCF1XnnWyA5sFgz"  # 账号里面的地址
    secret_key = "EDwRggaMuOn8LK3Ys59zaw2tVrysY0CA"  # 同上
    expired = time.time() + 2592000
    onceExpired = 0
    current = time.time()
    rdm = ''.join(random.choice("0123456789") for i in range(10))
    userid = "0"
    fileid = "tencentyunSignTest"

    info = "a=" + appid + "&k=" + secret_id + "&e=" + str(expired) + "&t=" + str(current) + "&r=" + str(
        rdm) + "&u=0&f="  # 去掉bucket

    signindex = hmac.new(secret_key, info, hashlib.sha1).digest()  # HMAC-SHA1加密
    sign = base64.b64encode(signindex + info)  # base64转码

    url = "http://recognition.image.myqcloud.com/ocr/general"
    headers = {'Host': 'recognition.image.myqcloud.com',
               "Authorization": sign,
               }
    files = {'appid': (None, appid),
             #  'bucket': (None, bucket),
             'image': ('yanzhengma.png', open('yanzhengma.png', 'rb'), 'image/jpeg')

             }

    r = requests.post(url, files=files, headers=headers)
    responseinfo = r.content

    r_index = r'itemstring":"(.*?)"'  # 做一个正则匹配
    result = re.findall(r_index, responseinfo)[0]
   # result=result[0]
    
    new_result = filter(str.isalnum,result)
    return new_result
	
def ks_login():
    
    #登录网址和信息
    url="https://investorservice.cfmmc.com/login.do"
    user_name='005280028009'
    passwd = 'xb7208dx'

    s = requests.Session()
    headers = {"User-Agent":"Mozilla/5.0"}
    r = s.get(url)
    value = re.findall(r'name="org.apache.struts.taglib.html.TOKEN" value="(.*?)"',r.text)[0]
    captcha = s.get("https://investorservice.cfmmc.com/veriCode.do?t={}".format(str(int(time.time()))), headers=headers)
    with open("captcha.png","wb") as f:
        f.write(captcha.content)
    #获取验证码
    vericode = yanzheng_local()
    print "六位数的验证码:{}".format(vericode)
    data = {
        "org.apache.struts.taglib.html.TOKEN":value,
        "showSaveCookies":"",
        "userID":user_name,
        "password":passwd,
        "vericode":vericode,
    }
    login = s.post("https://investorservice.cfmmc.com/login.do", headers=headers, data=data)
   
    aim_url="https://investorservice.cfmmc.com/customer/setupViewCustomerDetailFromCompanyAuto.do"
    new_url=s.get(aim_url)
     #返回值是一个respond的对象
    excel_url="https://investorservice.cfmmc.com/customer/setupViewCustomerDetailFromCompanyWithExcel.do"
    try:
        excel = s.get(excel_url)
        with open('cat.xls','wb') as f:
            f.write(excel.content)
    except:
        print"验证码没有识别，重新登录。。。"
        ks_login()

def ks_save():
	#连接数据库，存储到kunshan表中
    conn = pymysql.connect(
        host ='localhost',
        user='root',
        passwd = 'kunshan19',
        db='mis',
        port=3306,
        charset='utf8'
    )

    # 获得游标
    cur = conn.cursor()
    
    filename='cat.xls'
    book = xlrd.open_workbook(filename)
    sheet = book.sheets()[0]

    srjc= sheet.cell(11,2).value        #上日结存
    tradeday = sheet.cell(4, 7).value       #交易日期
    quanyi = sheet.cell(11,7).value       #客户权益
    code = sheet.cell(12, 2).value      #当日存取合计
    #syhb= sheet.cell(12, 7).value      #实有货币基金
    pcyk = sheet.cell(13, 2).value      #平仓盈亏
    fhbc = sheet.cell(13, 7).value      #非货币冲抵金额
    drzq = sheet.cell(14, 2).value      #当日总权利金
    hbcd = sheet.cell(14, 7).value      #货币充抵金额
    drsx = sheet.cell(15, 2).value      #当日手续费
    drjc = sheet.cell(16, 2).value      #当日结存
    bzjz = sheet.cell(16, 7).value      #保证金占用
    fdyk = sheet.cell(17, 2).value      # 浮动盈亏
    kyzj = sheet.cell(17, 7).value      #可用资金
    fxd = sheet.cell(18, 7).value       #风险度
    zjbz = sheet.cell(19, 7).value      #追加保证金
    #读取数据库昨日净值
    cur.execute("select jingzhi from ( select * from basedata_document order by id desc)as tmp")
    jz_before = cur.fetchone()[0]       #上一日的净值

    #前一天的当日存取金（表中最近的一条数据）
    cur.execute("select code from ( select * from basedata_document order by id desc)as tmp")
    cqhj_before = cur.fetchone()[0]     #上一日的存取合计
    #print type(cqhj_before)
    #份额
    cur.execute("select fene from ( select * from basedata_document order by id desc)as tmp")
    fene_before = cur.fetchone()[0]     #上一日的份额

    #今日份额，读取的字符串型数据转回float,进行计算
    fene=float(fene_before)+float(float(cqhj_before)/float(jz_before))   #今日的份额=昨日份额+昨日入金/昨日份额
	
    # 今日的净值=（客户权益-当日存取）/今日的份额
    jingzhi=(quanyi-code)/fene

    #插入数据，小数强制转换成字符串
    cur.execute("INSERT INTO basedata_document(tradeday,srjc,quanyi,code,pcyk,drzq,drsx,drjc,bzjz,fdyk,fxd,zjbz,fene,jingzhi) VALUES('"+str(tradeday)+"','"+str(srjc)+"','"+str(quanyi)+"','"+str(code)+"','"+str(pcyk)+"',\
        '"+str(drzq)+"','"+str(drsx)+"','"+str(drjc)+"','"+str(bzjz)+"','"+str(fdyk)+"','"+str(fxd)+"','"+str(zjbz)+"','"+str(fene)+"','"+str(jingzhi)+"')")

    conn.commit()
    print "存入{}号的数据到数据库完成".format(tradeday)

def main():
    #自动登录 下载数据
    ks_login()
    time.sleep(5)
	#存入数据库
    ks_save()
if __name__ == '__main__':
    main()
